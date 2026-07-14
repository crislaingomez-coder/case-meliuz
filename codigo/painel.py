from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from codigo.analisador import AnalysisResult


VERDE_MELIUZ = "426F49"
VERDE_ESCURO = "31583A"
ROSA_MELIUZ = "F48DB8"
ROSA_ESCURO = "D95F95"
ROSA_CLARO = "FCE4EF"
VERDE_CLARO = "E7F1E7"
CINZA_FUNDO = "F5F7F5"
CINZA_BORDA = "D8E3D8"
CINZA_TEXTO = "34423A"
BRANCO = "FFFFFF"
AMARELO_CLARO = "FFF0C7"
PALETA = [ROSA_MELIUZ, VERDE_MELIUZ, ROSA_ESCURO, "8CA98D", "E8B7CC"]


def build_dashboard(results: list[AnalysisResult], output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    chart_dir = output_path.parent / "graficos"
    chart_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo Executivo"

    _build_summary_sheet(ws, results, chart_dir)
    for result in results:
        _build_partner_sheet(wb, result, chart_dir)

    wb.save(output_path)
    return output_path


def _build_summary_sheet(ws, results: list[AnalysisResult], chart_dir: Path) -> None:
    _prepare_sheet(ws)
    _title(ws, "Dashboard de Testes A/B de Cashback", "A1:L2", "Analise automatizada de testes de cashback")

    winners = [
        result.metrics[result.metrics["grupo"] == result.decision_group].iloc[0]
        for result in results
    ]
    total_profit = sum(float(row["lucro_liquido_total"]) for row in winners)
    avg_margin = sum(float(row["margem_sobre_gmv"]) for row in winners) / len(winners)
    avg_roi = sum(float(row["roi_cashback"]) for row in winners) / len(winners)

    _kpi_card(ws, "A4:C6", "Testes analisados", len(results), "inteiro", ROSA_CLARO)
    _kpi_card(ws, "D4:F6", "Lucro vencedor total", total_profit, "moeda", VERDE_CLARO)
    _kpi_card(ws, "G4:I6", "Margem media", avg_margin, "percentual", AMARELO_CLARO)
    _kpi_card(ws, "J4:L6", "ROI medio", avg_roi, "percentual", ROSA_CLARO)

    headers = ["Parceiro", "Periodo", "Variante", "Lucro", "Margem", "ROI", "Decisao"]
    start_row = 9
    _write_header(ws, start_row, headers)
    for row_idx, result in enumerate(results, start=start_row + 1):
        winner = result.metrics[result.metrics["grupo"] == result.decision_group].iloc[0]
        values = [
            result.partner,
            result.period,
            result.decision_group,
            float(winner["lucro_liquido_total"]),
            float(winner["margem_sobre_gmv"]),
            float(winner["roi_cashback"]),
            result.decision,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _body_cell(cell)
        ws.row_dimensions[row_idx].height = 38
        ws.cell(row=row_idx, column=4).number_format = 'R$ #,##0.00'
        ws.cell(row=row_idx, column=5).number_format = '0.00%'
        ws.cell(row=row_idx, column=6).number_format = '0.00%'

    labels = [result.partner.replace("Parceiro ", "P") for result in results]
    profits = [float(row["lucro_liquido_total"]) for row in winners]
    rois = [float(row["roi_cashback"]) for row in winners]

    profit_img = _bar_chart_image(
        chart_dir / "resumo_lucro.png",
        "Lucro da variante recomendada",
        labels,
        profits,
        "moeda",
        [VERDE_MELIUZ, ROSA_MELIUZ, ROSA_ESCURO],
    )
    roi_img = _bar_chart_image(
        chart_dir / "resumo_roi.png",
        "ROI de cashback da variante recomendada",
        labels,
        rois,
        "percentual",
        [ROSA_MELIUZ, ROSA_MELIUZ, ROSA_MELIUZ],
    )
    _add_image(ws, profit_img, "A15", 570, 305)
    _add_image(ws, roi_img, "G15", 570, 305)

    _set_widths(ws, {"A": 17, "B": 22, "C": 13, "D": 15, "E": 12, "F": 12, "G": 40, "H": 12, "I": 12, "J": 12, "K": 12, "L": 12})
    ws.freeze_panes = None
    ws.sheet_view.zoomScale = 90


def _build_partner_sheet(wb: Workbook, result: AnalysisResult, chart_dir: Path) -> None:
    ws = wb.create_sheet(result.partner.replace("Parceiro ", "Parceiro_")[:31])
    _prepare_sheet(ws)
    _title(ws, result.test_name, "A1:N2", f"{result.period} | decisao automatizada")

    winner = result.metrics[result.metrics["grupo"] == result.decision_group].iloc[0]
    _kpi_card(ws, "A4:B6", "Variante recomendada", result.decision_group, "texto", VERDE_CLARO)
    _kpi_card(ws, "C4:D6", "Lucro liquido", float(winner["lucro_liquido_total"]), "moeda", VERDE_CLARO)
    _kpi_card(ws, "E4:F6", "Margem GMV", float(winner["margem_sobre_gmv"]), "percentual", AMARELO_CLARO)
    _kpi_card(ws, "G4:H6", "ROI cashback", float(winner["roi_cashback"]), "percentual", ROSA_CLARO)

    ws["A8"] = "Decisao"
    ws["A8"].font = Font(bold=True, color=VERDE_ESCURO)
    ws["B8"] = result.decision
    ws["B8"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("B8:H9")
    _fill_range(ws, "A8:H9", BRANCO)

    metrics = result.metrics.sort_values("grupo").reset_index(drop=True)
    table_start = 12
    headers = ["Grupo", "Compradores", "GMV", "Cashback", "Lucro", "Margem", "ROI", "Ticket medio"]
    _write_header(ws, table_start, headers)
    for idx, row in metrics.iterrows():
        excel_row = table_start + 1 + idx
        values = [
            row["grupo"],
            int(row["compradores"]),
            float(row["gmv_total"]),
            float(row["cashback_total"]),
            float(row["lucro_liquido_total"]),
            float(row["margem_sobre_gmv"]),
            float(row["roi_cashback"]),
            float(row["ticket_medio"]),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            _body_cell(cell)
        if row["grupo"] == result.decision_group:
            _fill_row(ws, excel_row, 1, len(headers), VERDE_CLARO)
        for col in [3, 4, 5, 8]:
            ws.cell(row=excel_row, column=col).number_format = 'R$ #,##0.00'
        for col in [6, 7]:
            ws.cell(row=excel_row, column=col).number_format = '0.00%'

    labels = metrics["grupo"].tolist()
    colors = [VERDE_MELIUZ if group == result.decision_group else ROSA_MELIUZ for group in labels]
    lucro_img = _bar_chart_image(
        chart_dir / f"{_slug(result.partner)}_lucro.png",
        "Lucro liquido por grupo",
        labels,
        metrics["lucro_liquido_total"].astype(float).tolist(),
        "moeda",
        colors,
    )
    roi_img = _bar_chart_image(
        chart_dir / f"{_slug(result.partner)}_roi.png",
        "ROI de cashback por grupo",
        labels,
        metrics["roi_cashback"].astype(float).tolist(),
        "percentual",
        colors,
    )
    _add_image(ws, lucro_img, "J4", 520, 285)
    _add_image(ws, roi_img, "J18", 520, 285)

    _daily_audit_table(ws, result)
    _set_widths(
        ws,
        {
            "A": 16,
            "B": 14,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 12,
            "G": 12,
            "H": 15,
            "I": 3,
            "J": 15,
            "K": 15,
            "L": 15,
            "M": 15,
            "N": 15,
            "O": 14,
            "P": 13,
            "Q": 13,
            "R": 13,
            "S": 13,
        },
    )
    ws.freeze_panes = "O1"
    ws.sheet_view.zoomScale = 90


def _daily_audit_table(ws, result: AnalysisResult) -> None:
    start_row = 3
    start_col = 15
    ws.cell(row=1, column=start_col, value="Dados diarios para auditoria").font = Font(
        bold=True, size=13, color=VERDE_ESCURO
    )
    daily = result.daily_metrics.sort_values(["data", "grupo"])
    pivot = (
        daily.pivot_table(index="data", columns="grupo", values="lucro_liquido", aggfunc="sum")
        .sort_index()
        .reset_index()
    )
    pivot["data"] = pd.to_datetime(pivot["data"]).dt.date.astype(str)
    headers = ["Data", *[str(col) for col in pivot.columns if col != "data"]]

    for offset, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + offset, value=header)
        cell.font = Font(bold=True, color=BRANCO)
        cell.fill = PatternFill("solid", fgColor=VERDE_MELIUZ)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()

    for row_idx, row in enumerate(pivot.itertuples(index=False), start=start_row + 1):
        for col_offset, value in enumerate(row):
            cell = ws.cell(row=row_idx, column=start_col + col_offset, value=value)
            _body_cell(cell)
            if col_offset > 0:
                cell.number_format = 'R$ #,##0.00'


def _bar_chart_image(
    path: Path,
    title: str,
    labels: list[str],
    values: list[float],
    value_kind: str,
    colors: list[str],
) -> Path:
    width, height = 980, 520
    margin_left, margin_right = 90, 52
    margin_top, margin_bottom = 86, 86
    chart_w = width - margin_left - margin_right
    chart_h = height - margin_top - margin_bottom
    image = Image.new("RGB", (width, height), "#FFFFFF")
    draw = ImageDraw.Draw(image)
    title_font = _font(30, bold=True)
    label_font = _font(20)
    value_font = _font(19, bold=True)

    draw.rounded_rectangle((8, 8, width - 8, height - 8), radius=28, outline="#E6D5DE", width=2, fill="#FFFFFF")
    draw.rectangle((8, 8, 20, height - 8), fill=f"#{ROSA_MELIUZ}")
    draw.text((margin_left, 28), title, fill=f"#{VERDE_ESCURO}", font=title_font)

    max_value = max(values) if values else 1
    max_value = max(max_value, 1)
    bar_gap = 52
    bar_w = max(70, int((chart_w - bar_gap * (len(values) - 1)) / max(len(values), 1)))
    x = margin_left
    baseline = margin_top + chart_h

    for idx, (label, value) in enumerate(zip(labels, values)):
        color = f"#{colors[idx % len(colors)]}"
        bar_h = int((value / max_value) * (chart_h * 0.82))
        y0 = baseline - bar_h
        x0 = x + idx * (bar_w + bar_gap)
        x1 = x0 + bar_w
        draw.rounded_rectangle((x0, y0, x1, baseline), radius=10, fill=color)
        draw.text((x0 + bar_w / 2, baseline + 18), label, fill=f"#{CINZA_TEXTO}", font=label_font, anchor="ma")
        draw.text((x0 + bar_w / 2, y0 - 30), _format_value(value, value_kind), fill=f"#{VERDE_ESCURO}", font=value_font, anchor="ma")

    draw.line((margin_left, baseline, width - margin_right, baseline), fill="#D8E3D8", width=2)
    path.parent.mkdir(parents=True, exist_ok=True)
    image.save(path)
    return path


def _add_image(ws, path: Path, anchor: str, width: int, height: int) -> None:
    image = ExcelImage(str(path))
    image.width = width
    image.height = height
    ws.add_image(image, anchor)


def _insert_logo(ws, anchor: str) -> None:
    logo_path = Path(__file__).resolve().parents[1] / "identidade_visual" / "logo-icon-2022-1080x1080.png"
    if not logo_path.exists():
        return
    image = ExcelImage(str(logo_path))
    image.width = 42
    image.height = 42
    ws.add_image(image, anchor)


def _prepare_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    _fill_range(ws, "A1:Z120", CINZA_FUNDO)
    for row in range(1, 121):
        ws.row_dimensions[row].height = 22


def _title(ws, text: str, cell_range: str, subtitle: str = "") -> None:
    start_cell = cell_range.split(":")[0]
    ws[start_cell] = text
    ws[start_cell].font = Font(bold=True, size=18, color=BRANCO)
    ws[start_cell].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(cell_range)
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=VERDE_MELIUZ)
    ws[start_cell] = f"     {text}\n     {subtitle}" if subtitle else f"     {text}"
    ws[start_cell].font = Font(bold=True, size=16, color=BRANCO)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 24
    _insert_logo(ws, "A1")


def _kpi_card(ws, cell_range: str, label: str, value, kind: str, fill: str) -> None:
    start, end = cell_range.split(":")
    ws.merge_cells(cell_range)
    ws[start] = f"{label}\n{_format_value(value, kind)}"
    ws[start].font = Font(bold=True, size=13, color=VERDE_ESCURO)
    ws[start].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    start_col = ws[start].column
    end_col = ws[end].column
    start_row = ws[start].row
    end_row = ws[end].row
    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].height = 27
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.border = _medium_border()


def _write_header(ws, row: int, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color=BRANCO)
        cell.fill = PatternFill("solid", fgColor=VERDE_MELIUZ)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()


def _body_cell(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=BRANCO)
    cell.border = _thin_border()
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.font = Font(color=CINZA_TEXTO)


def _fill_range(ws, cell_range: str, color: str) -> None:
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=color)


def _fill_row(ws, row: int, start_col: int, end_col: int, color: str) -> None:
    for col in range(start_col, end_col + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=color)


def _thin_border() -> Border:
    side = Side(style="thin", color=CINZA_BORDA)
    return Border(left=side, right=side, top=side, bottom=side)


def _medium_border() -> Border:
    side = Side(style="medium", color=BRANCO)
    return Border(left=side, right=side, top=side, bottom=side)


def _set_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for column_cells in ws.columns:
        col = get_column_letter(column_cells[0].column)
        if col not in widths:
            ws.column_dimensions[col].width = 12


def _format_value(value, kind: str) -> str:
    if kind == "moeda":
        return f"R$ {float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    if kind == "percentual":
        return f"{float(value) * 100:.2f}%".replace(".", ",")
    return str(value)


def _font(size: int, bold: bool = False):
    candidates = [
        "C:/Windows/Fonts/segoeuib.ttf" if bold else "C:/Windows/Fonts/segoeui.ttf",
        "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
    ]
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size=size)
        except OSError:
            continue
    return ImageFont.load_default()


def _slug(value: str) -> str:
    return value.lower().replace(" ", "_")
